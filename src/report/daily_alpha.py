"""Daily Alpha Report (Sprint 4.4).

Render each day's ranked Alpha candidates (already scored + persisted by the
4.2 ``ConsensusEngine``) into three interchangeable, date-archived formats:

* ``daily_alpha.xlsx`` — the data asset (openpyxl). Sheet1 = the candidate
  table (design §7 / v2 Sheet1); Sheet2 = a decision-tracking template with the
  system columns pre-filled and the human columns left blank (the human loop
  lands in 4.5).
* ``daily_alpha.html`` — the daily browser / WorkBuddy view; self-contained
  inline CSS + an SVG AROS bar chart, renders fully offline.
* ``daily_alpha.md``   — AI / knowledge-base friendly; mirrors the HTML content.

The renderer is **pure**: it consumes a list of :class:`~research.models.DailyAlphaCandidate`
rows (or queries them by ``run_date``) and writes files. No network, no scoring
math, no look-ahead. This keeps the report offline-testable and lets the daily
run regenerate a report from the DB at any time.
"""

from __future__ import annotations

import json
from collections.abc import Callable, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from research.models import DailyAlphaCandidate, DailyScreening

# A cell getter maps a candidate + the report date to a display value.
CellFn = Callable[[DailyAlphaCandidate, date], Any]


def _json_list(raw: str | None) -> list[str]:
    """Parse a JSON list column (concepts / hit_strategies); tolerate bad input."""
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        return [str(x) for x in parsed] if isinstance(parsed, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _fmt(v: Any) -> str:
    """Format a cell value for spreadsheets / tables.

    ``None`` -> dash (genuinely missing); empty string -> blank (deliberately
    left for a human to fill, e.g. the decision-tracking template's human cols).
    """
    if v is None:
        return "-"
    if v == "":
        return ""
    if isinstance(v, float):
        return f"{v:.1f}"
    if isinstance(v, date):
        return v.isoformat()
    return str(v)


# --------------------------------------------------------------------------- #
# Sheet specifications (design §7, v2 Sheet1 / Sheet2)
# --------------------------------------------------------------------------- #
SHEET1: list[tuple[str, CellFn]] = [
    ("日期", lambda c, rd: rd),
    ("代码", lambda c, rd: c.code),
    ("名称", lambda c, rd: c.name),
    ("行业", lambda c, rd: c.industry),
    ("板块", lambda c, rd: c.sector),
    ("概念", lambda c, rd: " / ".join(_json_list(c.concepts_json))),
    ("Regime", lambda c, rd: c.regime_label),
    ("命中套数", lambda c, rd: c.hit_count),
    ("命中策略", lambda c, rd: " / ".join(_json_list(c.hit_strategies_json))),
    ("平均星级", lambda c, rd: c.avg_quality_star),
    ("最高星级", lambda c, rd: c.max_quality_star),
    ("共振评分", lambda c, rd: c.consensus_score),
    ("公开资金", lambda c, rd: c.public_money_score),
    ("隐性行为", lambda c, rd: c.hidden_flow_score),
    ("板块强度", lambda c, rd: c.sector_score),
    ("AROS", lambda c, rd: c.aros_score),
    ("评级", lambda c, rd: c.rating),
    ("优势", lambda c, rd: c.advantages),
    ("风险", lambda c, rd: c.risks),
    ("Thesis", lambda c, rd: c.thesis),
    ("系统建议", lambda c, rd: c.system_suggestion),
    ("人工判断", lambda c, rd: ""),
    ("跟踪状态", lambda c, rd: ""),
]

SHEET2: list[tuple[str, CellFn]] = [
    ("候选日期", lambda c, rd: rd),
    ("代码·名称", lambda c, rd: f"{c.code} {c.name or ''}".strip()),
    ("系统评分·评级", lambda c, rd: f"{c.aros_score:.1f} {c.rating}"),
    ("人工决定", lambda c, rd: ""),
    ("人工理由", lambda c, rd: ""),
    ("计划·实际入场价", lambda c, rd: ""),
    ("计划·实际仓位", lambda c, rd: ""),
    ("复盘日期", lambda c, rd: ""),
    ("1·3·5·10日结果", lambda c, rd: ""),
    ("最大浮盈·浮亏", lambda c, rd: ""),
    ("最终收益", lambda c, rd: ""),
    ("是否验证系统", lambda c, rd: ""),
    ("复盘总结", lambda c, rd: ""),
]


def query_candidates(session: Session, run_date: date) -> list[DailyAlphaCandidate]:
    """Load a date's ranked candidates from the DB (design traceability chain).

    Joins ``daily_alpha_candidates`` to ``daily_screenings`` on ``run_date`` and
    returns them ordered by AROS score (desc), i.e. the Top-N the engine kept.
    """
    return list(
        session.query(DailyAlphaCandidate)
        .join(DailyScreening, DailyAlphaCandidate.screening_id == DailyScreening.id)
        .filter(DailyScreening.run_date == run_date)
        .order_by(DailyAlphaCandidate.aros_score.desc())
        .all()
    )


class DailyAlphaReport:
    """Render Daily Alpha candidates to Excel + HTML + Markdown, archived by date."""

    def generate(
        self,
        candidates: Sequence[DailyAlphaCandidate],
        run_date: date,
        out_dir: str | Path = "reports",
    ) -> dict[str, Path]:
        """Write all three formats under ``<out_dir>/<run_date>/``.

        Returns the absolute paths of the three files. Idempotent / re-runnable.
        """
        base = Path(out_dir) / str(run_date)
        base.mkdir(parents=True, exist_ok=True)
        xlsx = base / "daily_alpha.xlsx"
        html = base / "daily_alpha.html"
        md = base / "daily_alpha.md"

        self.write_excel(list(candidates), run_date, xlsx)
        html.write_text(self.to_html(list(candidates), run_date), encoding="utf-8")
        md.write_text(self.to_markdown(list(candidates), run_date), encoding="utf-8")
        return {"xlsx": xlsx.resolve(), "html": html.resolve(), "md": md.resolve()}

    # ------------------------------------------------------------------ #
    # Excel
    # ------------------------------------------------------------------ #
    def write_excel(
        self,
        candidates: list[DailyAlphaCandidate],
        run_date: date,
        path: str | Path,
    ) -> Path:
        path = Path(path)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Daily Alpha Candidate"
        self._fill_sheet(ws1, SHEET1, candidates, run_date)
        ws2 = wb.create_sheet("Decision Tracking")
        self._fill_sheet(ws2, SHEET2, candidates, run_date)
        wb.save(path)
        return path

    @staticmethod
    def _fill_sheet(
        ws: Any,
        spec: Sequence[tuple[str, CellFn]],
        candidates: list[DailyAlphaCandidate],
        run_date: date,
    ) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        ws.append([h for h, _ in spec])
        for col in range(1, len(spec) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in candidates:
            ws.append([_fmt(fn(c, run_date)) for _, fn in spec])
        for col in range(1, len(spec) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # HTML (self-contained, offline)
    # ------------------------------------------------------------------ #
    def to_html(self, candidates: list[DailyAlphaCandidate], run_date: date) -> str:
        from xml.sax.saxutils import escape

        def esc(s: Any) -> str:
            return escape(str(s)) if s is not None else ""

        gen = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        body: list[str] = []
        body.append("<h1>AROS 每日 Alpha 报告</h1>")
        body.append(
            f"<p class='meta'>生成时间: {esc(gen)} · 截面日期: {esc(run_date)} · "
            f"候选: {len(candidates)} 只</p>"
        )

        body.append("<h2>一、Top 候选</h2>")
        body.append(
            "<table><thead><tr>"
            "<th>排名</th><th>代码</th><th>名称</th><th>Regime</th>"
            "<th>命中</th><th>共振</th><th>AROS</th><th>评级</th>"
            "</tr></thead><tbody>"
        )
        for i, c in enumerate(candidates, 1):
            body.append(
                "<tr>"
                f"<td>{i}</td>"
                f"<td>{esc(c.code)}</td>"
                f"<td>{esc(c.name or '-')}</td>"
                f"<td>{esc(c.regime_label)}</td>"
                f"<td>{c.hit_count}</td>"
                f"<td>{c.consensus_score:.1f}</td>"
                f"<td>{c.aros_score:.1f}</td>"
                f"<td>{esc(c.rating)}</td>"
                "</tr>"
            )
        body.append("</tbody></table>")

        if candidates:
            body.append(self._aros_chart(candidates))

        body.append("<h2>二、候选明细</h2>")
        for i, c in enumerate(candidates, 1):
            body.append(self._candidate_detail(c, i))

        return (
            "<!DOCTYPE html><html lang='zh-CN'><head>"
            "<meta charset='utf-8'>"
            f"<title>AROS Alpha {esc(run_date)}</title>"
            f"<style>{_HTML_CSS}</style></head>"
            f"<body>{''.join(body)}</body></html>"
        )

    @staticmethod
    def _aros_chart(candidates: list[DailyAlphaCandidate]) -> str:
        scores = [c.aros_score for c in candidates]
        hi = max(scores)
        rows: list[str] = []
        for c, sc in zip(candidates, scores, strict=True):
            w = (sc / hi * 100.0) if hi else 0.0
            rows.append(
                f"<div class='bar-row'><span class='bar-label'>{c.code}</span>"
                f"<span class='bar-track'><span class='bar-fill' "
                f"style='width:{w:.1f}%'></span></span>"
                f"<span class='bar-val'>{sc:.1f}</span></div>"
            )
        return "<h2>三、AROS 评分分布</h2>" + "<div class='chart'>" + "".join(rows) + "</div>"

    @staticmethod
    def _candidate_detail(c: DailyAlphaCandidate, rank: int) -> str:
        from xml.sax.saxutils import escape

        def esc(s: Any) -> str:
            return escape(str(s)) if s is not None else ""

        block = [f"<div class='cand'><h3>#{rank} {esc(c.code)} {esc(c.name or '')}</h3>"]
        block.append(
            "<p class='tags'>"
            f"<span class='tag rating'>{esc(c.rating)}</span> "
            f"AROS <b>{c.aros_score:.1f}</b> · 共振 {c.consensus_score:.1f} · "
            f"Regime {esc(c.regime_label)} · 命中 {c.hit_count} 套"
            f"</p>"
        )
        if c.avg_quality_star is not None:
            max_star = f"{c.max_quality_star:.1f}" if c.max_quality_star is not None else "-"
            block.append(f"<p>平均星级 {c.avg_quality_star:.1f} · 最高星级 {max_star}</p>")
        for label, value in (
            ("优势", c.advantages),
            ("风险", c.risks),
            ("Thesis", c.thesis),
            ("系统建议", c.system_suggestion),
        ):
            if value:
                block.append(f"<p><b>{label}:</b> {esc(value)}</p>")
        block.append("</div>")
        return "".join(block)

    # ------------------------------------------------------------------ #
    # Markdown (AI / knowledge-base friendly)
    # ------------------------------------------------------------------ #
    def to_markdown(self, candidates: list[DailyAlphaCandidate], run_date: date) -> str:
        lines: list[str] = []
        lines.append("# AROS 每日 Alpha 报告")
        lines.append("")
        lines.append(f"- 截面日期: {run_date}")
        lines.append(f"- 候选数量: {len(candidates)} 只")
        lines.append("")
        if not candidates:
            lines.append("> 无候选标的（无信号或评分不足）。")
            return "\n".join(lines)

        lines.append("## 一、Top 候选")
        lines.append("")
        header = ["排名", "代码", "名称", "Regime", "命中", "共振", "AROS", "评级"]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("|" + "|".join(["---"] * len(header)) + "|")
        for i, c in enumerate(candidates, 1):
            lines.append(
                f"| {i} | {c.code} | {c.name or '-'} | {c.regime_label} | "
                f"{c.hit_count} | {c.consensus_score:.1f} | {c.aros_score:.1f} | {c.rating} |"
            )
        lines.append("")

        lines.append("## 二、候选明细")
        lines.append("")
        for i, c in enumerate(candidates, 1):
            lines.append(f"### #{i} {c.code} {c.name or ''}")
            lines.append("")
            lines.append(f"- 评级: {c.rating}")
            lines.append(f"- AROS: {c.aros_score:.1f} · 共振评分: {c.consensus_score:.1f}")
            lines.append(f"- Regime: {c.regime_label} · 命中: {c.hit_count} 套")
            if c.avg_quality_star is not None:
                max_star = f"{c.max_quality_star:.1f}" if c.max_quality_star is not None else "-"
                lines.append(f"- 平均星级: {c.avg_quality_star:.1f} · 最高星级: {max_star}")
            if c.advantages:
                lines.append(f"- 优势: {c.advantages}")
            if c.risks:
                lines.append(f"- 风险: {c.risks}")
            if c.thesis:
                lines.append(f"- Thesis: {c.thesis}")
            if c.system_suggestion:
                lines.append(f"- 系统建议: {c.system_suggestion}")
            lines.append("")
        return "\n".join(lines)


_HTML_CSS = """
:root { color-scheme: light; }
body { font-family: -apple-system, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
       margin: 24px; color: #1f2933; background: #fff; }
h1 { color: #1F4E78; border-bottom: 2px solid #1F4E78; padding-bottom: 8px; }
h2 { color: #1F4E78; margin-top: 28px; }
.meta { color: #52606d; font-size: 13px; }
table { border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 14px; }
th, td { border: 1px solid #d9e2ec; padding: 6px 10px; text-align: center; }
th { background: #1F4E78; color: #fff; }
tbody tr:nth-child(even) { background: #f5f7fa; }
.tag { display: inline-block; padding: 1px 8px; border-radius: 10px; font-size: 12px; }
.tag.rating { background: #1F4E78; color: #fff; }
.cand { border: 1px solid #d9e2ec; border-radius: 8px; padding: 12px 16px; margin: 12px 0; }
.cand h3 { margin: 0 0 6px; color: #1F4E78; }
.tags { color: #52606d; font-size: 13px; }
.chart { margin: 12px 0; }
.bar-row { display: flex; align-items: center; gap: 8px; margin: 4px 0; font-size: 13px; }
.bar-label { width: 70px; text-align: right; font-family: monospace; }
.bar-track { flex: 1; background: #e4e7eb; border-radius: 4px; height: 14px; overflow: hidden; }
.bar-fill { display: block; height: 100%; background: linear-gradient(90deg,#1F4E78,#3b82f6); }
.bar-val { width: 48px; text-align: right; font-family: monospace; }
"""
