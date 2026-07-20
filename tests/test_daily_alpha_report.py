"""Tests for the Sprint 4.4 Daily Alpha Report renderer (report/daily_alpha.py).

Fully offline: candidates are built in memory (or via an in-memory SQLite
round-trip with DailyScreening) and the three formats are asserted on disk.
No network, no scoring math.
"""

from __future__ import annotations

import json
from datetime import date

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from core.database import Base
from report.daily_alpha import (
    SHEET1,
    SHEET2,
    DailyAlphaReport,
    query_candidates,
)
from research.models import DailyAlphaCandidate, DailyScreening


def _mem_session():
    eng = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(eng)
    return sessionmaker(eng)()


def _candidate(code: str, **kw: object) -> DailyAlphaCandidate:
    return DailyAlphaCandidate(
        id=f"dac_{code}",
        screening_id=kw.get("screening_id", "scr_x"),
        code=code,
        name=kw.get("name", "测试股"),
        industry=kw.get("industry", "行业X"),
        sector=kw.get("sector", "板块X"),
        concepts_json=json.dumps(kw.get("concepts", ["概念A", "概念B"]), ensure_ascii=False),
        regime_label=kw.get("regime", "Bull"),
        hit_count=kw.get("hit_count", 3),
        hit_strategies_json=json.dumps(kw.get("hits", ["s1", "s2"]), ensure_ascii=False),
        avg_quality_star=kw.get("avg_star", 4.0),
        max_quality_star=kw.get("max_star", 5.0),
        consensus_score=kw.get("consensus", 80.0),
        public_money_score=kw.get("public", 70.0),
        hidden_flow_score=kw.get("hidden", 60.0),
        sector_score=kw.get("sector_score", 75.0),
        aros_score=kw.get("aros", 82.0),
        rating=kw.get("rating", "A"),
        advantages=kw.get("advantages", "多策略共振"),
        risks=kw.get("risks", "短期回撤"),
        thesis=kw.get("thesis", "thesis text"),
        system_suggestion=kw.get("suggestion", "系统建议"),
    )


def test_generate_writes_three_formats(tmp_path) -> None:
    cands = [_candidate("000001", aros=88.0), _candidate("600000", aros=72.0)]
    paths = DailyAlphaReport().generate(cands, date(2026, 7, 20), out_dir=tmp_path)
    for key in ("xlsx", "html", "md"):
        assert key in paths
        assert paths[key].exists()
        assert paths[key].stat().st_size > 0
    # archived under reports/<date>/
    assert paths["xlsx"].parent.name == "2026-07-20"


def test_excel_sheets_and_headers(tmp_path) -> None:
    cands = [_candidate("000001", name="平安银行", rating="A+", aros=90.0)]
    xlsx = DailyAlphaReport().generate(cands, date(2026, 7, 20), out_dir=tmp_path)["xlsx"]

    wb = load_workbook(xlsx)
    assert wb.sheetnames == ["Daily Alpha Candidate", "Decision Tracking"]

    ws1 = wb["Daily Alpha Candidate"]
    headers = [c.value for c in ws1[1]]
    assert headers == [h for h, _ in SHEET1]
    # data row carries code + rating + thesis
    row = [c.value for c in ws1[2]]
    assert "000001" in row
    assert "平安银行" in row
    assert "A+" in row
    assert "thesis text" in row
    # concepts joined by " / "
    assert "概念A / 概念B" in row

    ws2 = wb["Decision Tracking"]
    assert [c.value for c in ws2[1]] == [h for h, _ in SHEET2]
    # human columns left blank (system-prefilled only)
    human_blank = [c.value for c in ws2[2]]
    assert all(v in (None, "") for v in human_blank[3:])  # 人工决定..复盘总结 are blank
    # system side is prefilled
    assert "000001 平安银行" == ws2.cell(row=2, column=2).value
    assert "90.0 A+" == ws2.cell(row=2, column=3).value


def test_markdown_contains_key_fields(tmp_path) -> None:
    cands = [_candidate("000001", rating="A", aros=82.0, thesis="共振选股")]
    md = DailyAlphaReport().generate(cands, date(2026, 7, 20), out_dir=tmp_path)["md"]
    text = md.read_text(encoding="utf-8")
    assert "AROS 每日 Alpha 报告" in text
    assert "000001" in text
    assert "A" in text
    assert "共振选股" in text
    assert "82.0" in text


def test_html_contains_chart_and_detail(tmp_path) -> None:
    cands = [_candidate("000001", rating="A", aros=82.0, advantages="多策略共振")]
    html = DailyAlphaReport().generate(cands, date(2026, 7, 20), out_dir=tmp_path)["html"]
    text = html.read_text(encoding="utf-8")
    assert text.startswith("<!DOCTYPE html>")
    assert "AROS 每日 Alpha 报告" in text
    assert "000001" in text
    assert "多策略共振" in text
    assert "bar-fill" in text  # SVG bar chart rendered


def test_query_candidates_roundtrip_orders_by_aros() -> None:
    session = _mem_session()
    session.add(
        DailyScreening(
            id="scr1",
            run_date=date(2026, 7, 20),
            universe="csi800",
            regime_label="Bull",
        )
    )
    session.add(_candidate("000001", screening_id="scr1", aros=90.0))
    session.add(_candidate("600000", screening_id="scr1", aros=70.0))
    session.commit()

    rows = query_candidates(session, date(2026, 7, 20))
    assert [r.code for r in rows] == ["000001", "600000"]  # AROS desc
    # wrong date -> empty
    assert query_candidates(session, date(2026, 7, 19)) == []


def test_report_from_db_roundtrip(tmp_path) -> None:
    session = _mem_session()
    session.add(
        DailyScreening(
            id="scr2",
            run_date=date(2026, 7, 20),
            universe="csi800",
            regime_label="Bull",
        )
    )
    session.add(_candidate("000001", screening_id="scr2", name="平安银行", aros=85.0))
    session.commit()

    cands = query_candidates(session, date(2026, 7, 20))
    paths = DailyAlphaReport().generate(cands, date(2026, 7, 20), out_dir=tmp_path)
    assert paths["xlsx"].exists()
    text = paths["md"].read_text(encoding="utf-8")
    assert "平安银行" in text
